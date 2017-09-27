from openpyxl import Workbook
import subprocess
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,PatternFill,Border,Side
import time
import email
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.utils import COMMASPACE, formatdate
from email import encoders
from email.mime.application import MIMEApplication

"""
This script is expecting the ServiceNow files to be CSV and the fields to be in this order
name, dns_domain, exclusion, sys_class, operation, serial_number, location,managed_by,assigned_to,support_group,owned_by,ip_addy,asset_tag,sys_created,install_status,validation,model_number

This function takes the ServiceNow file and creates a dictionary with the name as the key
and everything else as a value
"""

def getServiceNowIPsAndNames(list):
    global ServiceNowNodes
    global nodesNotResolveInDNS
    global nodesIPDoesNotPing
    global nodesNameDoesPing
    global totalTentSystems
    global totalEntSystems
    totalEntSystems=0
    totalTentSystems=0
    serviceNowDict={}
    with open(list) as f:
        next(f)
        for line in f:
            line=line.split(',')
            for x in range(len(line)):
                line[x]=line[x].strip('"\n')
            if(line[0] not in serviceNowDict):
                serviceNowDict[line[0]]=[]
            for x in range(1,15):
                serviceNowDict[line[0]].append(line[x])
    for item in serviceNowDict:
        if(serviceNowDict[item][0]=="ent.rt.csaa.com" and serviceNowDict[item][1]!="Yes" and serviceNowDict[item][1]!="Hold"):
            totalEntSystems+=1
        elif(serviceNowDict[item][0]=="tent.trt.csaa.pri" and serviceNowDict[item][1]!="Yes" and serviceNowDict[item][1]!="Hold"):
            totalTentSystems+=1
    ServiceNowNodes=len(serviceNowDict)
    nodesNotResolveInDNS=0
    nodesIPDoesNotPing=0
    nodesNameDoesPing=0
    return serviceNowDict


"""
Fields should be in this order
    Device Name	IP Address	Device Category	Device Class | Sub-class	DID	Organization	Current State	Collection Group	Collection State	SNMP Credential	SNMP Version
The first field should be blank. So column A is all blank

This function takes the sciencelogic file and creates a dictionary with the name as the key
and everything else as a value
"""

def getScienceLogicIPsandNames(SLlist):
    global nodesInScienceLogic
    global nodesNotInScienceLogic
    global nodesNameDoesNotPing
    global nodesNameDoesPing
    global entSystems
    global tentSystems
    entSystems=0
    tentSystems=0
    scienceLogicDict={}
    with open(SLlist) as f:
        next(f)
        for line in f:
            line=line.split(',')
            for x in range(len(line)):
                line[x]=line[x].strip('"\n')
            if(line[1] not in scienceLogicDict):
                scienceLogicDict[line[1]]=[]
            for x in range(2,7):
                scienceLogicDict[line[1]].append(line[x])
    nodesInScienceLogic=0
    nodesNotInScienceLogic=0
    nodesNameDoesNotPing=0
    nodesNaMeDoesPing=0
    return scienceLogicDict

"""
This function resolves the dns by calling to the os system and running nslookup. If "Name" is in the
output, it's been resolved. If not, it hasn't.
"""
def resolveDNS(name):
    #Create globals to call in other functions
    global dnsResolve
    global nodesNotResolveInDNS
    output= subprocess.check_output(['nslookup', name])
    if "Name" in str(output):
        dnsResolve="Yes"
    else:
        nodesNotResolveInDNS+=1
        dnsResolve="No"

"""
This function pings the server. If it doesn't get an error, the server was pinged. if it does,
then it hasn't or timed out.
"""
def pingServer(ipaddy):
    global doesPing
    global nodesIPDoesNotPing
    global nodesNameDoesNotPing
    #Try statement, if it fails, it doesn't ping
    try:
        output= subprocess.check_output(["ping","-n","1", ipaddy])
        if (ipaddy!=""):
            doesPing="Yes"
        else:
            nodesIPDoesNotPing+=1
            doesPing="No"
    except subprocess.CalledProcessError:
        nodesIPDoesNotPing+=1
        doesPing="No"

#Function to count how many can be pinged by name
def pingServerName(name):
    global nodesNameDoesNotPing
    global nodesNameDoesPing
    global doesNamePing
    try:
        output= subprocess.check_output(["ping","-n","1", name])
        if (name!=""):
            nodesNameDoesPing+=1
            doesNamePing="Yes"
        else:
            nodesNameDoesNotPing+=1
            doesNamePing="No"
    except subprocess.CalledProcessError:
        nodesNameDoesNotPing+=1
        doesNamePing="No"

#check to see if the ips are in sciencelogic
def inScienceLogic(dict,ip,name,exclusion,system):
    global ServiceNowNodes
    global nodesInScienceLogic
    global nodesNotInScienceLogic
    global entSystems
    global tentSystems
    global doesNamePing
    global doesPing
    global dnsResolve
    doesPing="Not_Checked"
    dnsResolve="Not_Checked"
    doesNamePing="Not_Checked"
    if(exclusion=="Yes" or exclusion=="Hold"):
        ServiceNowNodes-=1
        return "Not_Checked"
    elif (name in dict and name!=""):
        nodesInScienceLogic+=1
        if(system=="ent.rt.csaa.com"):
            entSystems+=1
        if(system=="tent.trt.csaa.pri"):
            tentSystems+=1
        return "Yes"
    #Need it to change so if the ip address is blank it skips this.
    elif (ip in [x for v in dict.values() for x in v] and ip != ""):
        nodesInScienceLogic+=1
        if(system=="ent.rt.csaa.com"):
            entSystems+=1
        if(system=="tent.trt.csaa.pri"):
            tentSystems+=1
        return "Yes"
    else:
        nodesNotInScienceLogic+=1
        resolveDNS(name)
        pingServer(ip)
        pingServerName(name)
        return "No"

"""
This function will create the entire excel book and near the end call another function to make
a separate sheet for the totals
"""
def createExcel(list,ScienceLogicDict,ServiceNowDict):
    global pingConnects
    pingConnects=0
    #Create book with initial columns
    RESULTS =[['Name','DNS Domain','Exclusion','Class','Location','Managed'
    ,'Assigned to','Support Group','IP Address','Is the node found in Sciencelogic','Device Name','IP Address',
    'Device Class','If not found in Sciencelogic does the name resolve in DNS?'
    ,'If not found does the node ping by IP','Does the node ping by name']]
    wb = Workbook()
    ws = wb['Sheet']
    for i in range(len(RESULTS)):
        ws.append(RESULTS[i])
    #Fill the sheet with first columns. The last three columns start with a call to see if it's in
    #sciencelogic, if it's not call two other functions to see about dns and ping. Those functions
    #have global variables that are put into the excel sheet.
    for item in ServiceNowDict:
        #ws.append(name,domain[0],exclusion[1],class[2],location[5],managed[6],assigned[7],support[8],ip[10])
        name,domain,exclusion,serverClass,location,managed,assigned,support,ip=(item,ServiceNowDict[item][0],
        ServiceNowDict[item][1],ServiceNowDict[item][2],ServiceNowDict[item][5],ServiceNowDict[item][6],
        ServiceNowDict[item][7],ServiceNowDict[item][8],ServiceNowDict[item][10])
        slName=""
        slIP=""
        slClass=""
        answer=inScienceLogic(ScienceLogicDict,ip,name,exclusion,domain)
        if(answer=="Yes"):
            for item in ScienceLogicDict:
                if(name==item or ip in ScienceLogicDict[item]):
                    slName=item
                    slIP=ScienceLogicDict[item][0]
                    slClass=ScienceLogicDict[item][2]
                    if(slClass=="Ping | ICMP"):
                        pingConnects+=1

        ws.append([name,domain,exclusion,serverClass,location,managed,assigned,support,ip,
        answer,slName,slIP,slClass,dnsResolve,doesPing,doesNamePing])
    adjustColWidth(ws)
    #Give color
    giveSheet1Color(ws)
    createMathResults(wb,ScienceLogicDict,list)
    wb.save(list+" Coverage Report.xlsx")

def giveSheet1Color(sheet):
    blueFill = PatternFill(start_color='00FFFF',
                   end_color='00FFFF',
                   fill_type='solid')
    salmonFill = PatternFill(start_color='FFA07A',
                   end_color='00FFFF',
                   fill_type='solid')
    yellowFill = PatternFill(start_color='FFFF00',
                   end_color='00FFFF',
                   fill_type='solid')
    purpleFill = PatternFill(start_color='8A2BE2',
                   end_color='00FFFF',
                   fill_type='solid')

    y=65
    while (y<74):
         sheet[chr(y)+str(1)].fill = blueFill
         sheet[chr(y)+str(1)].font = Font(color="000000")
         y+=1
    sheet[chr(y)+str(1)].fill = yellowFill
    y+=1
    while(y<78):
         sheet[chr(y)+str(1)].fill = purpleFill
         sheet[chr(y)+str(1)].font = Font(color="000000")
         y+=1
    while(y<81):
         sheet[chr(y)+str(1)].fill = salmonFill
         sheet[chr(y)+str(1)].font = Font(color="000000")
         y+=1

def adjustColWidth(sheet):
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    for col in sheet.columns:
        max_length = 6
        column = col[0].column # Get the column name
        for cell in col:
            cell.border=thin_border
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    return sheet


def createMathResults(wb,sLDict,list):
    ws1 = wb.create_sheet("Totals")
    global ServiceNowNodes
    global nodesInScienceLogic
    global nodesNotInScienceLogic
    global entSystems
    global tentSystems
    global totalEntSystems
    global totalTentSystems
    #Create initial columns
    RESULTS =['Total Nodes from ServiceNow Checked',
                'Total Nodes from ServiceNow that are found in ScienceLogic',
                '%Total Nodes from ServiceNow that are found in ScienceLogic',
                'Total Nodes from ServiceNow that are not found in ScienceLogic',
                '%Total Nodes from ServiceNow that are not found in ScienceLogic',
                'Items from ServiceNow where name does not resolve in DNS',
                'Items from ServiceNow where name does not ping',
                'Items from ServiceNow where IP does not ping',
                '% of ENT systems covered by ScienceLogic',
                '% of TENT systems covered by ScienceLogic',
                'Nodes that are ping only connectivity'
                ]
    print("totalEntSystems,entSystems,totalTentSystems,tentSystems")
    print(totalEntSystems,entSystems,totalTentSystems,tentSystems)
    input()
    if(totalEntSystems!=0):
        entSystems=round(entSystems/totalEntSystems,4)
    else:
        entSystems=0
    if(totalTentSystems!=0):
        tentSystems=round(tentSystems/totalTentSystems,4)
    else:
        tentSystems=0
    NUMBERS = [ServiceNowNodes,nodesInScienceLogic,round(nodesInScienceLogic/ServiceNowNodes,4),
    nodesNotInScienceLogic,round(nodesNotInScienceLogic/ServiceNowNodes,4),nodesNotResolveInDNS,
    nodesNameDoesNotPing,nodesIPDoesNotPing,entSystems,tentSystems,pingConnects]

    ws1['B3'].number_format='00.00%'
    ws1['B5'].number_format='00.00%'
    ws1['B9'].number_format='00.00%'
    ws1['B10'].number_format='00.00%'
    # Add results to initial columns
    y=65
    for i in range(len(RESULTS)):
        ws1[chr(y)+str(i+1)]=RESULTS[i]
        ws1[chr(y+1)+str(i+1)]=NUMBERS[i]

    adjustColWidth(ws1)
    y=65
    blueFill = PatternFill(start_color='00FFFF',
                   end_color='00FFFF',
                   fill_type='solid')
     #Give top a blue color
    ws1[chr(y)+str(1)].fill = blueFill
    ws1[chr(y)+str(2)].fill = blueFill
    ws1[chr(y)+str(3)].fill = blueFill
    ws1[chr(y)+str(4)].fill = blueFill
    ws1[chr(y)+str(5)].fill = blueFill
    ws1[chr(y)+str(6)].fill = blueFill
    ws1[chr(y)+str(7)].fill = blueFill
    ws1[chr(y)+str(8)].fill = blueFill
    ws1[chr(y)+str(9)].fill = blueFill
    ws1[chr(y)+str(10)].fill = blueFill
    ws1[chr(y)+str(11)].fill = blueFill
    y+=1


    wb.save(list+" Coverage Report.xlsx")

def sendEmail(list):
    global ServiceNowNodes
    global nodesInScienceLogic
    global nodesNotInScienceLogic
    #Set up email address to send from
    MY_ADDRESS = 'DLInfrastructureTools@ent.rt.csaa.com'
    # set up the SMTP server
    try:
        s = smtplib.SMTP('172.26.156.125', 25)
        s.starttls()
        s.login('user','pass')
    except:
        msg = MIMEMultipart()
        start=time.strftime("%H:%M:%S")
        date=time.strftime("%m/%d/%Y")
        message="Ran the server coverage script on "+str(list) +". Started the script at " + str(start) + " on " + str(date) + "\t\n"
        message+=("# Total Nodes from ServiceNow Checked: " + str(ServiceNowNodes) + "\t\n" +
        "# Total Nodes from ServiceNow that are found in ScienceLogic: " + str(nodesInScienceLogic) + "\t\n" +
        "% Total Nodes from ServiceNow that are found in ScienceLogic: " + str(round(nodesInScienceLogic/ServiceNowNodes,2)) + "\t\n" +
        "# Total Nodes from ServiceNow that are not found in ScienceLogic: " + str(nodesNotInScienceLogic) +"\t\n" +
        "% Total Nodes from ServiceNow that are not found in ScienceLogic: " + str(round(nodesNotInScienceLogic/ServiceNowNodes,2)) + "\t\n" +
        "# Items from ServiceNow where name does not resolve in DNS: " + str(nodesNotResolveInDNS) + "\t\n" +
        "# Items from ServiceNow where name does not ping: " + str(nodesNameDoesNotPing) + "\t\n"+
        "# Items from ServiceNow where IP does not ping: " + str(nodesIPDoesNotPing))


        # setup the parameters of the message
        msg['From']=MY_ADDRESS
        msg['To']='Suzanne.Eden@csaa.com'
        msg['Subject']="This is TEST"
        part = MIMEBase('application', "octet-stream")
        filename = list+" Coverage Report.xlsx"
        part.set_payload(open(filename, "rb").read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=filename)
        msg.attach(part)
        # add in the message body
        msg.attach(MIMEText(message, 'plain'))


        # send the message via the server set up earlier.
        s.send_message(msg)
        del msg

    # Terminate the SMTP session and close the connection
        s.quit()

def main():
    start=time.strftime("%H:%M:%S")
    answer=input("Do you want to send this output as an email to Suzanne? (Y/N)")
    print("The script analyzes ~80 machines in 60 seconds")
    SLlist=input("Please enter the Sciencelogic file name including the file extension (EX:sciencelogic.csv): ")
    list=input("Please enter the list name including the file extension (EX:WindowsList.csv): ")
    ServiceNowDict=getServiceNowIPsAndNames(list)
    ScienceLogicDict=getScienceLogicIPsandNames(SLlist)
    createExcel(list,ScienceLogicDict,ServiceNowDict)
    print("Finished " + list)
    print("Start time:",start,"End time:",time.strftime("%H:%M:%S"))
    if (answer=="Y"):
        sendEmail(list)





if __name__ == '__main__':
    main()
